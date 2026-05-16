import io
import os
import sqlite3
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'snacktrack-dev-secret-2024')

DATABASE = os.environ.get('DATABASE_PATH', 'products.db')


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_db() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS product_master (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                default_unit TEXT NOT NULL,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                master_id INTEGER,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                batch_number TEXT,
                location TEXT NOT NULL,
                mfg_date DATE,
                packing_date DATE,
                expiry_date DATE NOT NULL,
                quantity REAL NOT NULL,
                unit TEXT NOT NULL,
                notes TEXT,
                is_split INTEGER DEFAULT 0,
                status TEXT DEFAULT 'active',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (master_id) REFERENCES product_master(id)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                product_name TEXT NOT NULL,
                batch_number TEXT,
                from_location TEXT NOT NULL,
                to_location TEXT NOT NULL,
                quantity_moved REAL NOT NULL,
                unit TEXT NOT NULL,
                moved_date DATE NOT NULL,
                moved_by TEXT,
                reason TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (product_id) REFERENCES products(id)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL
            )
        ''')
        conn.commit()


def migrate_db():
    """Add any columns introduced after initial schema creation."""
    with get_db() as conn:
        existing = [r[1] for r in conn.execute('PRAGMA table_info(products)').fetchall()]
        if 'master_id' not in existing:
            conn.execute('ALTER TABLE products ADD COLUMN master_id INTEGER REFERENCES product_master(id)')
        if 'is_split' not in existing:
            conn.execute('ALTER TABLE products ADD COLUMN is_split INTEGER DEFAULT 0')
        if 'status' not in existing:
            conn.execute("ALTER TABLE products ADD COLUMN status TEXT DEFAULT 'active'")

        conn.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL
            )
        ''')

        # Create default admin user if no users exist
        count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
        if count == 0:
            default_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
            conn.execute(
                'INSERT INTO users (username, password_hash) VALUES (?, ?)',
                ('admin', generate_password_hash(default_password))
            )
        conn.commit()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated



@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        if user and check_password_hash(user['password_hash'], password):
            session.permanent = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            return redirect(url_for('index'))
        error = 'Invalid username or password.'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    error = None
    if request.method == 'POST':
        current = request.form['current_password']
        new_pw = request.form['new_password']
        confirm = request.form['confirm_password']
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not check_password_hash(user['password_hash'], current):
            error = 'Current password is incorrect.'
        elif new_pw != confirm:
            error = 'New passwords do not match.'
        elif len(new_pw) < 6:
            error = 'Password must be at least 6 characters.'
        else:
            db.execute('UPDATE users SET password_hash = ? WHERE id = ?',
                       (generate_password_hash(new_pw), session['user_id']))
            db.commit()
            flash('Password changed successfully.', 'success')
            return redirect(url_for('index'))
    return render_template('change_password.html', error=error)


@app.template_filter('days_left')
def days_left_filter(expiry_date_str):
    if not expiry_date_str:
        return None
    try:
        expiry = datetime.strptime(str(expiry_date_str), '%Y-%m-%d').date()
        return (expiry - date.today()).days
    except Exception:
        return None


@app.template_filter('expiry_status')
def expiry_status_filter(expiry_date_str):
    days = days_left_filter(expiry_date_str)
    if days is None:
        return 'unknown'
    if days < 0:
        return 'expired'
    elif days <= 7:
        return 'critical'
    elif days <= 30:
        return 'warning'
    return 'ok'


@app.template_filter('fmt_date')
def fmt_date_filter(date_str):
    if not date_str:
        return '—'
    try:
        return datetime.strptime(str(date_str), '%Y-%m-%d').strftime('%d %b %Y')
    except Exception:
        return date_str


@app.route('/')
@login_required
def index():
    db = get_db()
    today = date.today()
    week = today + timedelta(days=7)
    month = today + timedelta(days=30)

    total = db.execute('SELECT COUNT(*) FROM product_master').fetchone()[0]
    expired_count = db.execute(
        'SELECT COUNT(*) FROM products WHERE expiry_date < ?', (today,)
    ).fetchone()[0]
    critical_count = db.execute(
        'SELECT COUNT(*) FROM products WHERE expiry_date >= ? AND expiry_date <= ?', (today, week)
    ).fetchone()[0]
    warning_count = db.execute(
        'SELECT COUNT(*) FROM products WHERE expiry_date > ? AND expiry_date <= ?', (week, month)
    ).fetchone()[0]

    stock = db.execute('''
        SELECT
            COALESCE(SUM(CASE WHEN p.is_split=0 AND p.status='active' THEN p.quantity + COALESCE(m.total_moved,0) ELSE 0 END), 0) AS total_stock,
            COALESCE(SUM(CASE WHEN p.is_split=0 AND p.status='active' THEN COALESCE(m.total_moved,0) ELSE 0 END), 0)              AS moved_stock,
            COALESCE(SUM(CASE WHEN p.is_split=0 AND p.status='active' THEN p.quantity ELSE 0 END), 0)                             AS current_stock
        FROM products p
        LEFT JOIN (
            SELECT product_id, SUM(quantity_moved) AS total_moved
            FROM movements GROUP BY product_id
        ) m ON p.id = m.product_id
    ''').fetchone()
    completed_count = db.execute(
        "SELECT COUNT(*) FROM products WHERE status='completed'"
    ).fetchone()[0]

    recent = db.execute(
        'SELECT * FROM products ORDER BY expiry_date ASC LIMIT 10'
    ).fetchall()

    return render_template('index.html',
                           total=total,
                           expired_count=expired_count,
                           critical_count=critical_count,
                           warning_count=warning_count,
                           total_stock=stock['total_stock'],
                           moved_stock=stock['moved_stock'],
                           current_stock=stock['current_stock'],
                           completed_count=completed_count,
                           recent=recent)


@app.route('/products')
@login_required
def products():
    db = get_db()
    today = date.today()
    week = today + timedelta(days=7)
    month = today + timedelta(days=30)

    search = request.args.get('search', '').strip()
    category = request.args.get('category', '')
    location = request.args.get('location', '')
    status = request.args.get('status', '')

    query = '''
        SELECT p.*, COALESCE(m.total_moved, 0) AS total_moved
        FROM products p
        LEFT JOIN (
            SELECT product_id, SUM(quantity_moved) AS total_moved
            FROM movements GROUP BY product_id
        ) m ON p.id = m.product_id
        WHERE 1=1
    '''
    params = []

    if search:
        query += ' AND (p.name LIKE ? OR p.batch_number LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    if category:
        query += ' AND p.category = ?'
        params.append(category)
    if location:
        query += ' AND p.location LIKE ?'
        params.append(f'%{location}%')
    if status == 'completed':
        query += " AND p.status = 'completed'"
    else:
        query += " AND p.status = 'active'"
        if status == 'expired':
            query += ' AND p.expiry_date < ?'
            params.append(today)
        elif status == 'critical':
            query += ' AND p.expiry_date >= ? AND p.expiry_date <= ?'
            params.extend([today, week])
        elif status == 'warning':
            query += ' AND p.expiry_date > ? AND p.expiry_date <= ?'
            params.extend([week, month])
        elif status == 'ok':
            query += ' AND p.expiry_date > ?'
            params.append(month)

    query += ' ORDER BY p.expiry_date ASC'
    all_products = db.execute(query, params).fetchall()
    original_count = sum(1 for p in all_products if not p['is_split'])
    locations = [r[0] for r in db.execute('SELECT DISTINCT location FROM products ORDER BY location').fetchall()]

    return render_template('products.html',
                           products=all_products,
                           original_count=original_count,
                           search=search,
                           category=category,
                           location=location,
                           status=status,
                           locations=locations)


@app.route('/product-master')
@login_required
def product_master():
    db = get_db()
    masters = db.execute('''
        SELECT pm.*, COUNT(p.id) AS stock_count
        FROM product_master pm
        LEFT JOIN products p ON p.master_id = pm.id
        GROUP BY pm.id ORDER BY pm.name
    ''').fetchall()
    return render_template('product_master.html', masters=masters)


@app.route('/product-master/add', methods=['GET', 'POST'])
@login_required
def add_master():
    if request.method == 'POST':
        db = get_db()
        db.execute(
            'INSERT INTO product_master (name, category, default_unit, description) VALUES (?, ?, ?, ?)',
            (
                request.form['name'].strip(),
                request.form['category'],
                request.form['default_unit'],
                request.form.get('description', '').strip() or None,
            )
        )
        db.commit()
        flash('Product added to master list.', 'success')
        return redirect(url_for('product_master'))
    return render_template('master_form.html', master=None, action='Add')


@app.route('/product-master/<int:mid>/edit', methods=['GET', 'POST'])
@login_required
def edit_master(mid):
    db = get_db()
    master = db.execute('SELECT * FROM product_master WHERE id = ?', (mid,)).fetchone()
    if not master:
        flash('Product not found.', 'danger')
        return redirect(url_for('product_master'))
    if request.method == 'POST':
        db.execute(
            'UPDATE product_master SET name=?, category=?, default_unit=?, description=? WHERE id=?',
            (
                request.form['name'].strip(),
                request.form['category'],
                request.form['default_unit'],
                request.form.get('description', '').strip() or None,
                mid,
            )
        )
        db.commit()
        flash('Product master updated.', 'success')
        return redirect(url_for('product_master'))
    return render_template('master_form.html', master=master, action='Edit')


@app.route('/product-master/<int:mid>/delete', methods=['POST'])
@login_required
def delete_master(mid):
    db = get_db()
    in_use = db.execute('SELECT COUNT(*) FROM products WHERE master_id = ?', (mid,)).fetchone()[0]
    if in_use:
        flash('Cannot delete — this product has stock entries linked to it.', 'danger')
    else:
        db.execute('DELETE FROM product_master WHERE id = ?', (mid,))
        db.commit()
        flash('Product deleted from master list.', 'success')
    return redirect(url_for('product_master'))


@app.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product():
    db = get_db()
    if request.method == 'POST':
        master_id = request.form.get('master_id') or None
        if master_id:
            master_id = int(master_id)
        db.execute('''
            INSERT INTO products (master_id, name, category, batch_number, location, mfg_date,
                                  packing_date, expiry_date, quantity, unit, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            master_id,
            request.form['name'],
            request.form['category'],
            request.form.get('batch_number') or None,
            request.form['location'],
            request.form.get('mfg_date') or None,
            request.form.get('packing_date') or None,
            request.form['expiry_date'],
            float(request.form['quantity']),
            request.form['unit'],
            request.form.get('notes') or None,
        ))
        db.commit()
        flash('Stock entry added successfully!', 'success')
        return redirect(url_for('products'))

    masters = db.execute('SELECT * FROM product_master ORDER BY name').fetchall()
    preselect_master_id = request.args.get('master_id', type=int)
    return render_template('product_form.html', product=None, action='Add', masters=masters,
                           preselect_master_id=preselect_master_id)


@app.route('/products/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
def edit_product(pid):
    db = get_db()
    product = db.execute('SELECT * FROM products WHERE id = ?', (pid,)).fetchone()
    if not product:
        flash('Product not found.', 'danger')
        return redirect(url_for('products'))

    if request.method == 'POST':
        db.execute('''
            UPDATE products SET name=?, category=?, batch_number=?, location=?, mfg_date=?,
                packing_date=?, expiry_date=?, quantity=?, unit=?, notes=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        ''', (
            request.form['name'],
            request.form['category'],
            request.form.get('batch_number') or None,
            request.form['location'],
            request.form.get('mfg_date') or None,
            request.form.get('packing_date') or None,
            request.form['expiry_date'],
            float(request.form['quantity']),
            request.form['unit'],
            request.form.get('notes') or None,
            pid,
        ))
        db.commit()
        flash('Product updated successfully!', 'success')
        return redirect(url_for('products'))

    masters = db.execute('SELECT * FROM product_master ORDER BY name').fetchall()
    return render_template('product_form.html', product=product, action='Edit', masters=masters)


@app.route('/products/<int:pid>/delete', methods=['POST'])
@login_required
def delete_product(pid):
    db = get_db()
    db.execute('DELETE FROM products WHERE id = ?', (pid,))
    db.commit()
    flash('Product deleted.', 'success')
    return redirect(url_for('products'))


@app.route('/products/<int:pid>/complete', methods=['POST'])
@login_required
def complete_product(pid):
    db = get_db()
    product = db.execute('SELECT * FROM products WHERE id = ?', (pid,)).fetchone()
    if not product:
        flash('Product not found.', 'danger')
        return redirect(url_for('products'))
    db.execute(
        "UPDATE products SET status='completed', updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (pid,)
    )
    db.commit()
    flash(f'"{product["name"]}" marked as sold/completed.', 'success')
    return redirect(url_for('products'))


@app.route('/products/export')
@login_required
def export_products():
    db = get_db()
    products = db.execute('''
        SELECT p.*, COALESCE(m.total_moved, 0) AS total_moved
        FROM products p
        LEFT JOIN (
            SELECT product_id, SUM(quantity_moved) AS total_moved
            FROM movements GROUP BY product_id
        ) m ON p.id = m.product_id
        WHERE p.status = 'active' AND p.is_split = 0
        ORDER BY p.expiry_date ASC
    ''').fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'

    header_fill = PatternFill(start_color='1F7A4B', end_color='1F7A4B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['#', 'Product Name', 'Category', 'Batch No.', 'Location',
               'Mfg Date', 'Packing Date', 'Expiry Date',
               'Total Stock', 'Moved', 'Current Stock', 'Unit', 'Status', 'Notes']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    today = date.today()
    for i, p in enumerate(products, 1):
        days_left = (datetime.strptime(str(p['expiry_date']), '%Y-%m-%d').date() - today).days
        if days_left < 0:
            status = 'Expired'
        elif days_left <= 7:
            status = 'Critical'
        elif days_left <= 30:
            status = 'Warning'
        else:
            status = 'Safe'

        ws.append([
            i,
            p['name'],
            p['category'],
            p['batch_number'] or '',
            p['location'],
            p['mfg_date'] or '',
            p['packing_date'] or '',
            p['expiry_date'],
            round(p['quantity'] + p['total_moved'], 2),
            round(p['total_moved'], 2),
            round(p['quantity'], 2),
            p['unit'],
            status,
            p['notes'] or '',
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'products_{date.today().isoformat()}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/alerts')
@login_required
def alerts():
    db = get_db()
    today = date.today()
    week = today + timedelta(days=7)
    month = today + timedelta(days=30)

    expired = db.execute(
        'SELECT * FROM products WHERE expiry_date < ? ORDER BY expiry_date ASC', (today,)
    ).fetchall()
    critical = db.execute(
        'SELECT * FROM products WHERE expiry_date >= ? AND expiry_date <= ? ORDER BY expiry_date ASC',
        (today, week)
    ).fetchall()
    warning = db.execute(
        'SELECT * FROM products WHERE expiry_date > ? AND expiry_date <= ? ORDER BY expiry_date ASC',
        (week, month)
    ).fetchall()

    return render_template('alerts.html', expired=expired, critical=critical, warning=warning)


@app.route('/products/<int:pid>/move', methods=['GET', 'POST'])
@login_required
def move_product(pid):
    db = get_db()
    product = db.execute('SELECT * FROM products WHERE id = ?', (pid,)).fetchone()
    if not product:
        flash('Product not found.', 'danger')
        return redirect(url_for('products'))

    if product['quantity'] <= 0:
        flash(f'Cannot move "{product["name"]}" — stock is zero.', 'danger')
        return redirect(url_for('products'))

    if request.method == 'POST':
        to_location = request.form['to_location'].strip()
        qty_moved = float(request.form['quantity_moved'])
        moved_date = request.form['moved_date']
        moved_by = request.form.get('moved_by', '').strip() or None
        reason = request.form.get('reason', '').strip() or None

        if qty_moved <= 0 or qty_moved > product['quantity']:
            flash(f'Invalid quantity. Available stock: {product["quantity"]} {product["unit"]}.', 'danger')
            return render_template('move_product.html', product=product, today=date.today().isoformat())

        db.execute('''
            INSERT INTO movements (product_id, product_name, batch_number, from_location,
                to_location, quantity_moved, unit, moved_date, moved_by, reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            pid, product['name'], product['batch_number'],
            product['location'], to_location,
            qty_moved, product['unit'], moved_date, moved_by, reason
        ))

        remaining = round(product['quantity'] - qty_moved, 4)
        if remaining <= 0:
            # Full move: just update location
            db.execute(
                'UPDATE products SET location=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                (to_location, pid)
            )
        else:
            # Partial move: reduce original stock, create new entry at destination
            db.execute(
                'UPDATE products SET quantity=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                (remaining, pid)
            )
            db.execute('''
                INSERT INTO products (name, category, batch_number, location, mfg_date,
                    packing_date, expiry_date, quantity, unit, notes, is_split)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', (
                product['name'], product['category'], product['batch_number'],
                to_location, product['mfg_date'], product['packing_date'],
                product['expiry_date'], qty_moved, product['unit'], product['notes']
            ))

        db.commit()
        flash(f'"{product["name"]}" moved to {to_location} successfully!', 'success')
        return redirect(url_for('movements'))

    return render_template('move_product.html', product=product, today=date.today().isoformat())


@app.route('/movements/new', methods=['GET', 'POST'])
@login_required
def new_movement():
    db = get_db()
    all_products = db.execute(
        "SELECT * FROM products WHERE status='active' AND quantity > 0 ORDER BY name"
    ).fetchall()

    if request.method == 'POST':
        pid = int(request.form['product_id'])
        product = db.execute('SELECT * FROM products WHERE id = ?', (pid,)).fetchone()
        if not product:
            flash('Product not found.', 'danger')
            return redirect(url_for('new_movement'))

        if product['quantity'] <= 0:
            flash(f'Cannot move "{product["name"]}" — stock is zero.', 'danger')
            return redirect(url_for('new_movement'))

        to_location = request.form['to_location'].strip()
        qty_moved = float(request.form['quantity_moved'])

        if qty_moved <= 0 or qty_moved > product['quantity']:
            flash(f'Invalid quantity. Available stock: {product["quantity"]} {product["unit"]}.', 'danger')
            return redirect(url_for('new_movement'))
        moved_date = request.form['moved_date']
        moved_by = request.form.get('moved_by', '').strip() or None
        reason = request.form.get('reason', '').strip() or None

        db.execute('''
            INSERT INTO movements (product_id, product_name, batch_number, from_location,
                to_location, quantity_moved, unit, moved_date, moved_by, reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            pid, product['name'], product['batch_number'],
            product['location'], to_location,
            qty_moved, product['unit'], moved_date, moved_by, reason
        ))

        remaining = round(product['quantity'] - qty_moved, 4)
        if remaining <= 0:
            db.execute(
                'UPDATE products SET location=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                (to_location, pid)
            )
        else:
            db.execute(
                'UPDATE products SET quantity=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                (remaining, pid)
            )
            db.execute('''
                INSERT INTO products (name, category, batch_number, location, mfg_date,
                    packing_date, expiry_date, quantity, unit, notes, is_split)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', (
                product['name'], product['category'], product['batch_number'],
                to_location, product['mfg_date'], product['packing_date'],
                product['expiry_date'], qty_moved, product['unit'], product['notes']
            ))

        db.commit()
        flash(f'"{product["name"]}" moved to {to_location} successfully!', 'success')
        return redirect(url_for('movements'))

    return render_template('new_movement.html', products=all_products, today=date.today().isoformat())


@app.route('/movements')
@login_required
def movements():
    db = get_db()
    search = request.args.get('search', '').strip()
    location = request.args.get('location', '').strip()
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    query = '''
        SELECT m.*, p.expiry_date FROM movements m
        LEFT JOIN products p ON m.product_id = p.id
        WHERE 1=1
    '''
    params = []

    if search:
        query += ' AND (m.product_name LIKE ? OR m.batch_number LIKE ? OR m.moved_by LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    if location:
        query += ' AND (m.from_location LIKE ? OR m.to_location LIKE ?)'
        params.extend([f'%{location}%', f'%{location}%'])
    if from_date:
        query += ' AND m.moved_date >= ?'
        params.append(from_date)
    if to_date:
        query += ' AND m.moved_date <= ?'
        params.append(to_date)

    query += ' ORDER BY m.created_at DESC'
    all_movements = db.execute(query, params).fetchall()

    locations = [r[0] for r in db.execute(
        'SELECT DISTINCT location FROM products ORDER BY location'
    ).fetchall()]
    total_moves = db.execute('SELECT COUNT(*) FROM movements').fetchone()[0]

    return render_template('movements.html',
                           movements=all_movements,
                           search=search,
                           location=location,
                           from_date=from_date,
                           to_date=to_date,
                           locations=locations,
                           total_moves=total_moves)


with app.app_context():
    init_db()
    migrate_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
